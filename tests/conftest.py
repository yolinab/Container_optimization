"""
conftest.py — Programmatic Excel fixture generators for the Container Optimizer test suite.

All Excel files are created in pytest's tmp_path so they are cleaned up automatically.
Column names match the real input_final.xlsx format so tests exercise real parsing logic.
"""
import pytest
import openpyxl


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_HEADERS = [
    "Barcode", "Article", "Productname",
    "Pallet and packing size",   # dimension string: "1,15x1,15x1,20"
    "Pallet size",               # type code: A2, C2, C3, NP
    "Total pallet in container",
    "Total pallets in row",
    "External Packaging Quantity",   # THE order-qty column for pallet rows
    "Total number of pallets",
    "Currency order", "Delivery condition", "PoD",
    "Item price FOB", "Total ordered amount",
]


def _dim(l_m: float, w_m: float, h_m: float) -> str:
    """Return a dimension string in metres with comma decimal, e.g. '1,15x0,77x1,20'."""
    def fmt(v):
        return f"{v:.2f}".replace(".", ",")
    return f"{fmt(l_m)}x{fmt(w_m)}x{fmt(h_m)}"


def _write_excel(path, rows: list[dict]) -> None:
    """
    Write an Excel workbook to *path* with the standard headers.
    Each dict in *rows* should have keys matching _HEADERS (missing keys → blank).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(_HEADERS)
    for row in rows:
        ws.append([row.get(h, None) for h in _HEADERS])
    wb.save(path)


# ---------------------------------------------------------------------------
# Pallet-row builders
# ---------------------------------------------------------------------------

def _a2_row(count: int, barcode: str = "1111", name: str = "Test A2") -> dict:
    """115×115×120 cm A2 pallet. pallets_per_block=8 (band 89-130, stack=2 → 2×120=240cm)."""
    return {
        "Barcode": barcode,
        "Productname": name,
        "Pallet and packing size": _dim(1.15, 1.15, 1.20),
        "Pallet size": "A2",
        "External Packaging Quantity": count,
    }


def _c2_row(count: int, barcode: str = "2222", name: str = "Test C2") -> dict:
    """115×77×120 cm C2 pallet. band 89-130, stack=2 → 240cm."""
    return {
        "Barcode": barcode,
        "Productname": name,
        "Pallet and packing size": _dim(1.15, 0.77, 1.20),
        "Pallet size": "C2",
        "External Packaging Quantity": count,
    }


def _tall_a2_row(count: int) -> dict:
    """115×115×145 cm A2 pallet. band >130, BUT 115×115|>130 not in block table → skipped."""
    return {
        "Barcode": "3333",
        "Productname": "Tall pallet",
        "Pallet and packing size": _dim(1.15, 1.15, 1.45),
        "Pallet size": "A2",
        "External Packaging Quantity": count,
    }


def _very_tall_c3_row(count: int) -> dict:
    """115×77×145 cm C3 pallet. band >130, stack=1 → 145 cm. Fits through 259cm door fine."""
    return {
        "Barcode": "4444",
        "Productname": "Very tall C3",
        "Pallet and packing size": _dim(1.15, 0.77, 1.45),
        "Pallet size": "C3",
        "External Packaging Quantity": count,
    }


def _overheight_c2_row(count: int) -> dict:
    """115×77×135 cm pallet. band >130, stack=1 → 135cm block. Fine for door."""
    return {
        "Barcode": "5555",
        "Productname": "Overheight C2",
        "Pallet and packing size": _dim(1.15, 0.77, 1.35),
        "Pallet size": "C2",
        "External Packaging Quantity": count,
    }


def _unknown_fp_row(count: int) -> dict:
    """100×100 cm footprint — not in any block type table."""
    return {
        "Barcode": "6666",
        "Productname": "Unknown footprint",
        "Pallet and packing size": _dim(1.00, 1.00, 0.80),
        "Pallet size": "A2",
        "External Packaging Quantity": count,
    }


def _np_row(count: int, l_m=0.46, w_m=0.46, h_m=0.87, name="Test NP box") -> dict:
    """NP (loose box) row."""
    return {
        "Barcode": "9999",
        "Productname": name,
        "Pallet and packing size": _dim(l_m, w_m, h_m),
        "Pallet size": "NP",
        "External Packaging Quantity": count,
        "Total number of pallets": count,
    }


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def excel_minimal_valid(tmp_path):
    """8 A2 pallets at 115×115×120cm — produces exactly 1 block, 1 container."""
    path = tmp_path / "minimal_valid.xlsx"
    _write_excel(path, [_a2_row(8)])
    return str(path)


@pytest.fixture
def excel_multiples_needed(tmp_path):
    """7 A2 pallets — 7 mod 8 ≠ 0, should trigger multiples error."""
    path = tmp_path / "multiples_needed.xlsx"
    _write_excel(path, [_a2_row(7)])
    return str(path)


@pytest.fixture
def excel_all_too_tall(tmp_path):
    """
    115×115×145cm pallets. These fall into the '>130' height band.
    115×115|>130 is NOT in the block type table (only 115x77|>130 exists).
    So 0 blocks are built → RuntimeError about no pallets parsed / no blocks.
    """
    path = tmp_path / "all_too_tall.xlsx"
    _write_excel(path, [_tall_a2_row(4)])
    return str(path)


@pytest.fixture
def excel_unknown_footprint(tmp_path):
    """100×100 cm pallets — unknown footprint, all skipped."""
    path = tmp_path / "unknown_fp.xlsx"
    _write_excel(path, [_unknown_fp_row(8)])
    return str(path)


@pytest.fixture
def excel_single_block(tmp_path):
    """
    Exactly 4 A2 pallets (115×115×120cm) = 1 block of 4.
    After packing container 1, N=0 remaining → loop exits cleanly.
    Also tests the N=1 block solver scenario if used in a 2-container sequence.
    """
    path = tmp_path / "single_block.xlsx"
    _write_excel(path, [_a2_row(4, name="Single block")])
    return str(path)


@pytest.fixture
def excel_np_only(tmp_path):
    """Only NP (loose box) rows — no pallet rows. Should produce 0 containers."""
    path = tmp_path / "np_only.xlsx"
    _write_excel(path, [_np_row(10)])
    return str(path)


@pytest.fixture
def excel_mixed(tmp_path):
    """A2 + C2 (exact multiples) + NP rows."""
    path = tmp_path / "mixed.xlsx"
    _write_excel(path, [
        _a2_row(8, barcode="1001", name="Product A"),
        _c2_row(4, barcode="2001", name="Product B"),
        _np_row(6, name="Box X"),
    ])
    return str(path)


@pytest.fixture
def excel_wrong_col_name(tmp_path):
    """Order qty in column 'Qty' — not in the fuzzy-match candidate list."""
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["Barcode", "Productname", "Pallet and packing size", "Pallet size", "Qty"]
    ws.append(headers)
    ws.append(["1111", "Test", _dim(1.15, 1.15, 1.20), "A2", 8])
    path = tmp_path / "wrong_col.xlsx"
    wb.save(path)
    return str(path)


@pytest.fixture
def excel_wrong_col_override(tmp_path):
    """Same as wrong_col_name but used to test count_col_override success path."""
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["Barcode", "Productname", "Pallet and packing size", "Pallet size", "MyQty"]
    ws.append(headers)
    ws.append(["1111", "Test", _dim(1.15, 1.15, 1.20), "A2", 8])
    path = tmp_path / "override_col.xlsx"
    wb.save(path)
    return str(path)


@pytest.fixture
def excel_two_containers(tmp_path):
    """
    Enough A2 pallets to fill 2 containers.
    Container fits ~10 blocks of 4 × 115cm + gaps = 10*(115+5)=1200cm ≈ 1203cm.
    So 10 blocks = 40 pallets per container → 80 pallets = 2 containers.
    Use 16 blocks (64 A2 pallets) to guarantee 2 containers.
    """
    path = tmp_path / "two_containers.xlsx"
    _write_excel(path, [_a2_row(64)])
    return str(path)


@pytest.fixture
def excel_dot_decimal(tmp_path):
    """Dimension string with dot decimal separator: '1.15x1.15x1.20'."""
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = _HEADERS
    ws.append(headers)
    row = {h: None for h in headers}
    row["Barcode"] = "7777"
    row["Productname"] = "Dot decimal"
    row["Pallet and packing size"] = "1.15x1.15x1.20"  # dots, not commas
    row["Pallet size"] = "A2"
    row["External Packaging Quantity"] = 8
    ws.append([row.get(h) for h in headers])
    path = tmp_path / "dot_decimal.xlsx"
    wb.save(path)
    return str(path)


@pytest.fixture
def excel_blank_header_rows(tmp_path):
    """Excel with 2 blank rows before the header — tests _detect_header_row."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([None] * len(_HEADERS))   # blank row 1
    ws.append([None] * len(_HEADERS))   # blank row 2
    ws.append(_HEADERS)                  # actual header row (index 2)
    row = {h: None for h in _HEADERS}
    row["Productname"] = "Blank header test"
    row["Pallet and packing size"] = _dim(1.15, 1.15, 1.20)
    row["Pallet size"] = "A2"
    row["External Packaging Quantity"] = 8
    ws.append([row.get(h) for h in _HEADERS])
    path = tmp_path / "blank_header.xlsx"
    wb.save(path)
    return str(path)
